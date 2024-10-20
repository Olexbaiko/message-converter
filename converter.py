import pandas as pd
from openpyxl import Workbook
import argparse

def process_file(file_path):
    # Read the input file (change to your XML or CSV)

    df = pd.read_excel(file_path)

    # Prepare a new workbook
    wb = Workbook()

    # Group by Frequency only
    grouped = df.groupby(['Частота'])

    # Process each group
    for frequency, group in grouped:
        # Create sheet name based on Frequency only
        sheet_name = f"{frequency}"
        
        # Add a new sheet to the workbook
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        ws.append(["Хто викликав", "Кого викликав", "К-сть переговорів", "Часовий проміжок"])
        
        # Group by "Who called" and "Who received"
        who_called_groups = group.groupby(["Хто викликав", "Кого викликав"])
        
        for (caller, receiver), sub_group in who_called_groups:
            # Calculate Communication Count
            communication_count = len(sub_group)
            
            # Find the time range (first and last message time)
            time_range = f"{sub_group['Час'].min()} - {sub_group['Час'].max()}"
            
            # Append the row to the sheet
            ws.append([caller, receiver, communication_count, time_range])

    # Remove the default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save the output Excel file
    output_filename = f"output_file_{file_path.split('/')[-1].split('.')[0]}.xlsx"
    wb.save(output_filename)
    print(f"File saved as {output_filename}")


if __name__ == "__main__":
    # Setup argument parser
    parser = argparse.ArgumentParser(description='Process an XML/CSV file and create an Excel report.')
    parser.add_argument('--file', help='The path to the input XML or CSV file')

    # Parse command-line arguments
    args = parser.parse_args()

    # Call the processing function with the provided filename
    process_file(args.file)